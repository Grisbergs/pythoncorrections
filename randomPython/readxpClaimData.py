import os
from openpyxl import load_workbook
from pathlib import Path
import csv

def get_all_files(directory):
    """Recursively retrieves all file paths in the specified directory."""
    directory_path = Path(directory)
    return [str(file) for file in directory_path.rglob('*') if file.is_file()]

# Set input folder and output CSV path
directory = r"\\QTSPRODFILES7\common\Kofax\Kapow_Output\XLS\AI Workbooks"
csv_filename = r"C:\C#Class\claim_data_outputClaimData_test.csv"
file_paths = get_all_files(directory)

header_written = False

for file in file_paths:
    try:
        workbook = load_workbook(file, data_only=True)
        if 'CLAIMDATA' not in workbook.sheetnames:
            print(f"Skipping {file}: 'CLAIMDATA' sheet not found.")
            continue
        sheet = workbook['CLAIMDATA']

        claim_data = {
            "ClaimNumber": sheet['B3'].value,
            "SC_DateOfSettlement": sheet['C3'].value,
            "SC_NadaAtDOL": sheet['D3'].value,
            "SC_PrimarySettlementAmt": sheet['E3'].value,
            "SB_ActualCashValue": sheet['F3'].value,
            "SB_BreakdownAudited": sheet['G3'].value,
            "SB_DateOfLoss": sheet['H3'].value,
            "SB_InsuranceDeductible": sheet['I3'].value,
            "SB_OtherDeductions": sheet['J3'].value,
            "SB_PriorDamage": sheet['K3'].value,
            "SB_SalesTax": sheet['L3'].value,
            "SB_SalesTaxRate": sheet['M3'].value,
            "SB_Salvage": sheet['N3'].value,
            "SB_SettlementAmount": sheet['O3'].value,
            "SB_StorageCharges": sheet['P3'].value,
            "SB_TitleFee": sheet['Q3'].value,
            "SB_TowingCharges": sheet['R3'].value,
            "IE_ActualCashValue": sheet['S3'].value,
            "IE_AvgMileagePerYear": sheet['T3'].value,
            "IE_DollarAmountToDeduct": sheet['U3'].value,
            "IE_EvaluationSource": sheet['V3'].value,
            "IE_MileageAtDOL": sheet['W3'].value,
            "IE_MissedOptions": sheet['X3'].value,
            "IE_PriorDamage": sheet['Y3'].value,
            "IE_StateVehicleEvaluatedAtDOL": sheet['Z3'].value,
            "GC_ClaimSubmittalTimeFrame": sheet['AA3'].value,
            "GC_ClaimSubmittedInTimeFrame": sheet['AB3'].value,
            "GC_DateOfSale": sheet['AC3'].value,
            "GC_DealerStateAtDOP": sheet['AD3'].value,
            "GC_DeductibleCovered": sheet['AE3'].value,
            "GC_GapFormNumber": sheet['AF3'].value,
            "GC_GapFormRevisionDate": sheet['AG3'].value,
            "GC_MaxFinancingTerms": sheet['AH3'].value,
            "GC_MaxLiabilityPercent": sheet['AI3'].value,
            "GC_MileageAtDOP": sheet['AJ3'].value,
            "LLC_AmountFinanced": sheet['AK3'].value,
            "LLC_APRRate": sheet['AL3'].value,
            "LLC_DateOfSale": sheet['AM3'].value,
            "LLC_DaysToFirstPayment": sheet['AN3'].value,
            "LLC_FirstPaymentDue": sheet['AO3'].value,
            "LLC_LoanFee": sheet['AP3'].value,
            "LLC_MonthlyPayment": sheet['AQ3'].value,
            "LLC_NewOrUsed": sheet['AR3'].value,
            "LLC_OverLoanQuestion": sheet['AS3'].value,
            "LLC_PaymentsPerYear": sheet['AT3'].value,
            "LLC_Terms": sheet['AU3'].value,
            "LLC_WarrantiesPurchased": sheet['AV3'].value,
            "PH_AmortBalanceAtDOL": sheet['AW3'].value,
            "PH_LastPaymentDate": sheet['AX3'].value,
            "PH_LoanBalance": sheet['AY3'].value,
            "PH_LoanFundingAmount": sheet['AZ3'].value,
            "PH_PaymentsDue": sheet['BA3'].value,
            "PH_PaymentsMade": sheet['BB3'].value,
            "CW_CreditLifeDisabilityRefund": sheet['BC3'].value,
            "CW_MaintenanceRefund": sheet['BD3'].value,
            "CW_TheftRefund": sheet['BE3'].value,
            "CW_TireWheelRefund": sheet['BF3'].value,
            "CW_VSCRefund": sheet['BG3'].value,
            "PR_CauseOfLoss": sheet['BH3'].value,
            "PR_DOL": sheet['BI3'].value,
            "PR_Exclusions": sheet['BJ3'].value,
            "MO_MissedOption": sheet['BK3'].value,
            "MO_MissedOptionsTotalAfterTax": sheet['BL3'].value,
            "MSRP_MaxLiability": sheet['BM3'].value,
            "MSRP_MSRPNADAValue": sheet['BN3'].value,
            "make": f"{sheet['BP3'].value or ''} {sheet['BO3'].value or ''}".strip(),
            "modyear": f"{sheet['BP3'].value or ''} {sheet['BO3'].value or ''}".strip(),
            "sg_ac_desc": sheet['BQ3'].value,
            "sg_con_vin": sheet['BR3'].value,
            "customername": sheet['BS3'].value,
            "firstdocdate": sheet['BT3'].value,
            "sg_con_carrier": sheet['BU3'].value,
            "lienholder": sheet['BV3'].value,
            "UnitNumber": sheet['BW3'].value,
            "primaryINS": sheet['BX3'].value,
            "CreatedOn": sheet['BY3'].value,
            "IE_SalesTaxRate": sheet['BZ3'].value,
            "LLC_CashPriceNada": sheet['CA3'].value,
            "LLC_TitleFee": sheet['CB3'].value,
            "LLC_CW_Maintenance": sheet['CC3'].value,
            "LLC_CW_VSC": sheet['CD3'].value,
            "LLC_CW_TireWheel": sheet['CE3'].value,
            "LLC_CW_LifeDisability": sheet['CF3'].value,
            "LLC_CW_Theft": sheet['CG3'].value,
            "Location": file
        }

        write_mode = 'a' if header_written else 'w'
        with open(csv_filename, mode=write_mode, newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=claim_data.keys())
            if not header_written:
                writer.writeheader()
                header_written = True
            writer.writerow(claim_data)

        print(f"Appended data from '{file}' to CSV.")

    except Exception as e:
        print(f"Error processing file '{file}': {e}")
