import os
from openpyxl import load_workbook
from pathlib import Path
import csv

def get_all_files(directory):
    """Recursively retrieves all file paths in the specified directory."""
    directory_path = Path(directory)
    return [str(file) for file in directory_path.rglob('*') if file.is_file()]

# Set input folder and output CSV path
directory = r"\\QTSPRODFILES7\common\Kofax\Kapow_Output\XLS\AI Workbooks"
csv_filename = r"C:\C#Class\claim_data_output_test2.csv"
file_paths = get_all_files(directory)

# Initialize flag for header write
header_written = False

for file in file_paths:
    try:
        workbook = load_workbook(file, data_only=True)
        if 'Write Up' not in workbook.sheetnames:
            print(f"Skipping {file}: 'Write Up' sheet not found.")
            continue
        sheet = workbook['Write Up']

        claim_data = {
            "ClaimNumber": sheet['B3'].value,
            "SC_DateOfSettlement": sheet['B5'].value,
            "SC_NadaAtDOL": sheet['E7'].value,
            "SC_PrimarySettlementAmt": sheet['C7'].value,
            "SB_ActualCashValue": sheet['E9'].value,
            "SB_BreakdownAudited": sheet['C10'].value,
            "SB_DateOfLoss": sheet['B4'].value,
            "SB_InsuranceDeductible": sheet['C11'].value,
            "SB_OtherDeductions": sheet['E14'].value,
            "SB_SalesTax": sheet['C12'].value,
            "SB_SalesTaxRate": sheet['C14'].value,
            "SB_Salvage": sheet['E11'].value,
            "SB_SettlementAmount": sheet['C9'].value,
            "SB_StorageCharges": sheet['E13'].value,
            "SB_TitleFee": sheet['C13'].value,
            "SB_TowingCharges": sheet['E12'].value,
            "IE_ActualCashValue": sheet['E17'].value,
            "IE_EvaluationSource": sheet['C16'].value,
            "IE_MileageAtDOL": sheet['C18'].value,
            "IE_MissedOptions": sheet['C19'].value,
            "IE_StateVehicleEvaluatedAtDOL": sheet['C17'].value,
            "GC_DateOfSale": sheet['E22'].value,
            "GC_DealerStateAtDOP": sheet['E24'].value,
            "GC_DeductibleCovered": sheet['C22'].value,
            "GC_GapFormNumber": sheet['C21'].value,
            "GC_GapFormRevisionDate": sheet['E21'].value,
            "GC_MaxFinancingTerms": sheet['E25'].value,
            "GC_MaxLiabilityPercent": sheet['C24'].value,
            "GC_MileageAtDOP": sheet['E23'].value,
            "LLC_AmountFinanced": sheet['C30'].value,
            "LLC_APRRate": sheet['C33'].value,
            "LLC_DateOfSale": sheet['E29'].value,
            "LLC_FirstPaymentDue": sheet['E30'].value,
            "LLC_LoanFee": sheet['E32'].value,
            "LLC_MonthlyPayment": sheet['C37'].value,
            "LLC_NewOrUsed": sheet['C29'].value,
            "LLC_Terms": sheet['C32'].value,
            "LLC_WarrantiesPurchased": sheet['E33'].value,
            "PH_LastPaymentDate": sheet['C40'].value,
            "PH_LoanFundingAmount": sheet['C36'].value,
            "PH_PaymentsMade": sheet['C37'].value,
            "CW_CreditLifeDisabilityRefund": sheet['E46'].value,
            "CW_MaintenanceRefund": sheet['E45'].value,
            "CW_TheftRefund": sheet['E47'].value,
            "CW_TireWheelRefund": sheet['E48'].value,
            "CW_VSCRefund": sheet['E44'].value,
            "PR_DOL": sheet['C50'].value,
            "PR_Exclusions": sheet['E52'].value,
            "MO_MissedOptionsTotalAfterTax": sheet['E19'].value,
            "make": sheet['D3'].value,
            "modyear": sheet['D3'].value,
            "sg_ac_desc": sheet['E3'].value,
            "sg_con_vin": sheet['F3'].value,
            "firstdocdate": sheet['D4'].value,
            "sg_con_carrier": sheet['C3'].value,
            "lienholder": sheet['F35'].value,
            "UnitNumber": sheet['F50'].value,
            "primaryINS": sheet['F16'].value,
            "Location": file
        }

        # Determine write mode
        write_mode = 'a' if header_written else 'w'

        # Write to CSV
        with open(csv_filename, mode=write_mode, newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=claim_data.keys())
            if not header_written:
                writer.writeheader()
                header_written = True
            writer.writerow(claim_data)

        print(f"Appended data from '{file}' to CSV.")

    except Exception as e:
        print(f"Error processing file '{file}': {e}")
